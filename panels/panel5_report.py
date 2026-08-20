"""Panel 5 — Report & Export"""
import streamlit as st
import numpy as np
import pandas as pd
import io
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


def render():
    st.markdown("<div class='panel-title'>📄 Panel 5 — Report & Export</div>", unsafe_allow_html=True)

    if not st.session_state.get("mc_results"):
        st.markdown("<div class='warning-box'>⚠️ Please complete the volumetric calculation in Panel 4 first.</div>", unsafe_allow_html=True)
        return

    mc = st.session_state["mc_results"]
    petro = st.session_state.get("petro_params", {})
    grv_data = st.session_state.get("grv_data", {})
    segy_info = st.session_state.get("segy_info", {})

    st.markdown("""
    <div class='info-box'>
    Export your complete reservoir analysis as a professional PDF report or Excel workbook.
    All results, parameters, and charts are included.
    </div>
    """, unsafe_allow_html=True)

    # ── Report Preview ───────────────────────────────────────────────────────
    st.markdown("### 📋 Report Preview")

    col1, col2 = st.columns([3, 1])
    with col2:
        project_name = st.text_input("Project Name", "Reservoir Study 1")
        author = st.text_input("Author", "Geoscientist")
        field_name = st.text_input("Field Name", "Field A")

    with col1:
        _render_report_preview(mc, petro, grv_data, segy_info, project_name, author, field_name)

    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Export Options ────────────────────────────────────────────────────────
    st.markdown("### 💾 Export Options")

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("""
        <div class='panel-card' style='text-align:center;'>
            <div style='font-size:2rem; margin-bottom:8px;'>📊</div>
            <div style='font-weight:600; color:#2a9bb0; margin-bottom:8px;'>Excel Workbook</div>
            <div style='font-size:0.82rem; color:#8aafc0; margin-bottom:16px;'>
            All results, parameters, and simulation data in structured sheets
            </div>
        </div>
        """, unsafe_allow_html=True)
        excel_data = _generate_excel(mc, petro, grv_data, segy_info, project_name)
        if excel_data:
            st.download_button(
                label="⬇️ Download Excel",
                data=excel_data,
                file_name=f"GeoVol_{project_name.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    with c2:
        st.markdown("""
        <div class='panel-card' style='text-align:center;'>
            <div style='font-size:2rem; margin-bottom:8px;'>📄</div>
            <div style='font-weight:600; color:#f0a500; margin-bottom:8px;'>PDF Report</div>
            <div style='font-size:0.82rem; color:#8aafc0; margin-bottom:16px;'>
            Professional report with results summary and key figures
            </div>
        </div>
        """, unsafe_allow_html=True)
        pdf_data = _generate_pdf(mc, petro, grv_data, segy_info, project_name, author, field_name)
        if pdf_data:
            st.download_button(
                label="⬇️ Download PDF",
                data=pdf_data,
                file_name=f"GeoVol_{project_name.replace(' ', '_')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        else:
            st.caption("Install reportlab: `pip install reportlab`")

    with c3:
        st.markdown("""
        <div class='panel-card' style='text-align:center;'>
            <div style='font-size:2rem; margin-bottom:8px;'>📋</div>
            <div style='font-weight:600; color:#2ecc71; margin-bottom:8px;'>CSV Data</div>
            <div style='font-size:0.82rem; color:#8aafc0; margin-bottom:16px;'>
            Raw Monte Carlo simulation data for external processing
            </div>
        </div>
        """, unsafe_allow_html=True)
        csv_data = _generate_csv(mc)
        st.download_button(
            label="⬇️ Download CSV",
            data=csv_data,
            file_name=f"GeoVol_{project_name.replace(' ', '_')}_MC.csv",
            mime="text/csv",
            use_container_width=True,
        )

    # ── Horizon Export ────────────────────────────────────────────────────────
    if st.session_state.get("horizon_top") is not None:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("### 🗺️ Export Horizon Grids")

        info = st.session_state.get("segy_info", {})
        sample_rate = info.get("sample_rate", 4.0)
        top_h = st.session_state["horizon_top"]
        base_h = st.session_state["horizon_base"]

        c1, c2 = st.columns(2)
        with c1:
            top_csv = _horizon_to_csv(top_h, sample_rate, "Top_Reservoir")
            st.download_button("⬇️ Top Horizon Grid (CSV)", top_csv,
                               "top_horizon.csv", "text/csv", use_container_width=True)
        with c2:
            base_csv = _horizon_to_csv(base_h, sample_rate, "Base_Reservoir")
            st.download_button("⬇️ Base Horizon Grid (CSV)", base_csv,
                               "base_horizon.csv", "text/csv", use_container_width=True)

        st.markdown("""
        <div class='info-box'>
        Horizon grids are exported in (Inline, Crossline, Time_ms) format —
        compatible with OpendTect, Petrel, and Kingdom import.
        </div>
        """, unsafe_allow_html=True)


def _render_report_preview(mc, petro, grv_data, segy_info, project, author, field):
    """Render an inline HTML report preview."""
    from datetime import datetime
    date_str = datetime.now().strftime("%B %d, %Y")

    results_html = ""
    for fluid, label, unit in [("oiip", "OIIP", "MMstb"), ("giip", "GIIP", "Bscf")]:
        if fluid in mc:
            r = mc[fluid]
            results_html += f"""
            <tr>
                <td style='padding:8px; border-bottom:1px solid #2a3a4a;'><b>{label}</b></td>
                <td style='padding:8px; border-bottom:1px solid #2a3a4a; color:#e74c3c;'>{r['p10']:.1f} {unit}</td>
                <td style='padding:8px; border-bottom:1px solid #2a3a4a; color:#2ecc71;'><b>{r['p50']:.1f} {unit}</b></td>
                <td style='padding:8px; border-bottom:1px solid #2a3a4a; color:#f39c12;'>{r['p90']:.1f} {unit}</td>
                <td style='padding:8px; border-bottom:1px solid #2a3a4a;'>{r['mean']:.1f} {unit}</td>
            </tr>"""

    petro_html = ""
    for key, label in [("phi", "Porosity"), ("sw", "Water Sat"), ("ntg", "NTG")]:
        if key in petro:
            p = petro[key]
            petro_html += f"""
            <tr>
                <td style='padding:6px; border-bottom:1px solid #2a3a4a;'>{label}</td>
                <td style='padding:6px; border-bottom:1px solid #2a3a4a;'>{p['p10']:.3f}</td>
                <td style='padding:6px; border-bottom:1px solid #2a3a4a;'>{p['p50']:.3f}</td>
                <td style='padding:6px; border-bottom:1px solid #2a3a4a;'>{p['p90']:.3f}</td>
                <td style='padding:6px; border-bottom:1px solid #2a3a4a;'>{p.get('source','—')}</td>
            </tr>"""

    grv_info = ""
    if grv_data:
        grv_info = f"GRV: {grv_data.get('grv_acre_ft',0):,.0f} acre-ft &nbsp;|&nbsp; Avg thickness: {grv_data.get('avg_thickness_m',0):.1f}m"

    st.markdown(f"""
    <div style='background:#0f2139; border:1px solid rgba(42,155,176,0.3); border-radius:12px;
                padding:28px; font-family: Space Grotesk, sans-serif;'>

        <!-- Header -->
        <div style='border-bottom:2px solid #2a9bb0; padding-bottom:16px; margin-bottom:20px;'>
            <div style='font-size:1.4rem; font-weight:700; color:#e8f4f8;'>🛢️ GeoVol — Reservoir Volumetrics Report</div>
            <div style='color:#8aafc0; margin-top:4px; font-size:0.85rem;'>
                Project: <b style='color:#e8f4f8;'>{project}</b> &nbsp;|&nbsp;
                Field: <b style='color:#e8f4f8;'>{field}</b> &nbsp;|&nbsp;
                Author: <b style='color:#e8f4f8;'>{author}</b> &nbsp;|&nbsp;
                Date: {date_str}
            </div>
        </div>

        <!-- Seismic Info -->
        <div style='font-size:0.82rem; color:#8aafc0; margin-bottom:16px;'>
            <b style='color:#2a9bb0;'>SEISMIC CUBE</b> &nbsp;|&nbsp;
            {segy_info.get('n_inlines','—')} inlines × {segy_info.get('n_crosslines','—')} crosslines &nbsp;|&nbsp;
            {segy_info.get('total_time_ms','—')} ms TWT &nbsp;|&nbsp; {grv_info}
        </div>

        <!-- Volumetric Results -->
        <div style='font-size:0.9rem; font-weight:600; color:#2a9bb0; margin-bottom:10px; text-transform:uppercase; letter-spacing:0.8px;'>
            Volumetric Results ({mc.get('n_simulations',0):,} Monte Carlo simulations)
        </div>
        <table style='width:100%; border-collapse:collapse; font-size:0.88rem; margin-bottom:20px;'>
            <thead>
                <tr style='background:#162840;'>
                    <th style='padding:8px; text-align:left; color:#8aafc0;'>Parameter</th>
                    <th style='padding:8px; text-align:left; color:#e74c3c;'>P10 (Low)</th>
                    <th style='padding:8px; text-align:left; color:#2ecc71;'>P50 (Best)</th>
                    <th style='padding:8px; text-align:left; color:#f39c12;'>P90 (High)</th>
                    <th style='padding:8px; text-align:left; color:#8aafc0;'>Mean</th>
                </tr>
            </thead>
            <tbody>{results_html}</tbody>
        </table>

        <!-- Petrophysics -->
        <div style='font-size:0.9rem; font-weight:600; color:#2a9bb0; margin-bottom:10px; text-transform:uppercase; letter-spacing:0.8px;'>
            Petrophysical Parameters
        </div>
        <table style='width:100%; border-collapse:collapse; font-size:0.85rem;'>
            <thead>
                <tr style='background:#162840;'>
                    <th style='padding:6px; text-align:left; color:#8aafc0;'>Parameter</th>
                    <th style='padding:6px; text-align:left; color:#8aafc0;'>P10</th>
                    <th style='padding:6px; text-align:left; color:#8aafc0;'>P50</th>
                    <th style='padding:6px; text-align:left; color:#8aafc0;'>P90</th>
                    <th style='padding:6px; text-align:left; color:#8aafc0;'>Source</th>
                </tr>
            </thead>
            <tbody>{petro_html}</tbody>
        </table>

        <div style='margin-top:20px; font-size:0.75rem; color:#4a6a7a; text-align:center; border-top:1px solid #2a3a4a; padding-top:12px;'>
            Generated by GeoVol Reservoir Volumetrics App &nbsp;|&nbsp;
            For exploration purposes only — verify against qualified petroleum engineer assessment
        </div>
    </div>
    """, unsafe_allow_html=True)


def _generate_excel(mc, petro, grv_data, segy_info, project_name):
    """Generate Excel workbook with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Volumetric Summary"

        header_fill = PatternFill("solid", fgColor="0D3D4A")
        subheader_fill = PatternFill("solid", fgColor="1A6B7A")
        result_fill = PatternFill("solid", fgColor="162840")
        header_font = Font(name="Calibri", bold=True, color="E8F4F8", size=11)
        subheader_font = Font(name="Calibri", bold=True, color="2A9BB0", size=10)

        ws["A1"] = f"GeoVol — Reservoir Volumetrics Report"
        ws["A1"].font = Font(name="Calibri", bold=True, color="F0A500", size=14)
        ws["A2"] = f"Project: {project_name}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(name="Calibri", color="8AAFC0", size=10)

        row = 4
        ws.cell(row, 1, "VOLUMETRIC RESULTS (Monte Carlo)").font = subheader_font

        row += 1
        headers = ["Parameter", "P10 (Low)", "P50 (Best Estimate)", "P90 (High)", "Mean", "Std Dev", "Unit", "Simulations"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h)
            c.font = header_font
            c.fill = header_fill

        row += 1
        for fluid, label, unit in [("oiip", "OIIP", "MMstb"), ("giip", "GIIP", "Bscf")]:
            if fluid in mc:
                r = mc[fluid]
                for col, val in enumerate([label, r["p10"], r["p50"], r["p90"], r["mean"], r["std"], unit, mc["n_simulations"]], 1):
                    c = ws.cell(row, col, val)
                    c.fill = result_fill
                    if col > 1 and col < 7:
                        c.number_format = "0.00"
                row += 1

        # Petrophysics section
        row += 2
        ws.cell(row, 1, "PETROPHYSICAL PARAMETERS").font = subheader_font
        row += 1
        for col, h in enumerate(["Parameter", "P10", "P50", "P90", "Source"], 1):
            c = ws.cell(row, col, h)
            c.font = header_font
            c.fill = header_fill

        row += 1
        for key, label in [("phi", "Porosity (φ)"), ("sw", "Water Sat (Sw)"),
                            ("ntg", "NTG"), ("bo", "Bo (rb/stb)"), ("bg", "Bg (rcf/scf)")]:
            if key in petro:
                p = petro[key]
                for col, val in enumerate([label, p["p10"], p["p50"], p["p90"], p.get("source","—")], 1):
                    ws.cell(row, col, val)
                row += 1

        # GRV section
        if grv_data:
            row += 2
            ws.cell(row, 1, "GRV CALCULATION").font = subheader_font
            row += 1
            for k, v in [("GRV (acre-ft)", grv_data.get("grv_acre_ft", 0)),
                         ("GRV (Mm³)", grv_data.get("grv_m3", 0) / 1e6),
                         ("Avg Thickness (m)", grv_data.get("avg_thickness_m", 0)),
                         ("Max Thickness (m)", grv_data.get("max_thickness_m", 0)),
                         ("Active Cells", grv_data.get("n_cells", 0))]:
                ws.cell(row, 1, k)
                ws.cell(row, 2, round(v, 2))
                row += 1

        # Column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 22

        # ── Sheet 2: MC Simulation Data ──────────────────────────────────────
        ws2 = wb.create_sheet("MC Simulation Data")
        ws2["A1"] = "Monte Carlo Raw Simulation Data (first 10,000 rows)"
        ws2["A1"].font = subheader_font

        col_headers = []
        col_data = []
        for fluid in ["oiip", "giip"]:
            if fluid in mc:
                col_headers.append(f"{fluid.upper()} ({mc[fluid]['unit']})")
                col_data.append(mc[fluid]["raw"][:10000])

        for col, h in enumerate(col_headers, 1):
            ws2.cell(2, col, h).font = header_font
        for row_i, vals in enumerate(zip(*col_data), 3):
            for col, v in enumerate(vals, 1):
                ws2.cell(row_i, col, round(float(v), 4))

        # ── Sheet 3: Sensitivity ─────────────────────────────────────────────
        ws3 = wb.create_sheet("Sensitivity Analysis")
        ws3["A1"] = "Sensitivity Analysis (Pearson Correlation Coefficients)"
        ws3["A1"].font = subheader_font

        row = 3
        for fluid in ["oiip", "giip"]:
            if fluid in mc.get("sensitivity", {}):
                ws3.cell(row, 1, f"{fluid.upper()} Sensitivity").font = header_font
                row += 1
                for param, corr in mc["sensitivity"][fluid].items():
                    ws3.cell(row, 1, param)
                    ws3.cell(row, 2, round(corr, 4))
                    row += 1
                row += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except ImportError:
        st.warning("Install openpyxl: `pip install openpyxl`")
        return None
    except Exception as e:
        st.error(f"Excel generation error: {e}")
        return None


def _generate_pdf(mc, petro, grv_data, segy_info, project_name, author, field_name):
    """Generate PDF report using reportlab."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, HRFlowable)
        from datetime import datetime

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()
        teal = colors.HexColor("#2a9bb0")
        dark = colors.HexColor("#0d3d4a")
        accent = colors.HexColor("#f0a500")
        light = colors.HexColor("#e8f4f8")
        subtle = colors.HexColor("#8aafc0")

        title_style = ParagraphStyle("Title", parent=styles["Title"],
                                      fontSize=18, textColor=accent, spaceAfter=4)
        sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                                    fontSize=10, textColor=subtle, spaceAfter=12)
        section_style = ParagraphStyle("Section", parent=styles["Heading2"],
                                        fontSize=12, textColor=teal, spaceBefore=16, spaceAfter=8)
        body_style = ParagraphStyle("Body", parent=styles["Normal"],
                                     fontSize=9, textColor=colors.HexColor("#cccccc"), spaceAfter=6)

        story = []

        # Title
        story.append(Paragraph("🛢️ GeoVol — Reservoir Volumetrics Report", title_style))
        story.append(Paragraph(
            f"Project: <b>{project_name}</b>  |  Field: <b>{field_name}</b>  |  "
            f"Author: <b>{author}</b>  |  Date: {datetime.now().strftime('%B %d, %Y')}",
            sub_style))
        story.append(HRFlowable(width="100%", thickness=1, color=teal, spaceAfter=16))

        # Seismic Info
        story.append(Paragraph("Seismic Cube Information", section_style))
        sei_data = [["Parameter", "Value"]]
        for k, v in [("Inlines", segy_info.get("n_inlines", "—")),
                     ("Crosslines", segy_info.get("n_crosslines", "—")),
                     ("Sample Rate", f"{segy_info.get('sample_rate','—')} ms"),
                     ("Total TWT", f"{segy_info.get('total_time_ms','—')} ms"),
                     ("Data Source", segy_info.get("source", "—").upper())]:
            sei_data.append([k, str(v)])

        t = Table(sei_data, colWidths=[8*cm, 9*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), teal),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#162840"), colors.HexColor("#0f2139")]),
            ("TEXTCOLOR", (0, 1), (-1, -1), light),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1a3a4a")),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t)

        # Volumetric Results
        story.append(Paragraph(f"Volumetric Results  ({mc.get('n_simulations',0):,} Monte Carlo Simulations)", section_style))
        vol_data = [["Parameter", "P10 (Low)", "P50 (Best)", "P90 (High)", "Mean", "Unit"]]
        for fluid, label, unit in [("oiip", "OIIP", "MMstb"), ("giip", "GIIP", "Bscf")]:
            if fluid in mc:
                r = mc[fluid]
                vol_data.append([label, f"{r['p10']:.2f}", f"{r['p50']:.2f}", f"{r['p90']:.2f}", f"{r['mean']:.2f}", unit])

        t2 = Table(vol_data, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 3*cm, 2.5*cm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), teal),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#162840"), colors.HexColor("#0f2139")]),
            ("TEXTCOLOR", (0, 1), (-1, -1), light),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1a3a4a")),
            ("PADDING", (0, 0), (-1, -1), 7),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        story.append(t2)

        # Petrophysics
        story.append(Paragraph("Petrophysical Parameters", section_style))
        pet_data = [["Parameter", "P10", "P50", "P90", "Source"]]
        for key, label in [("phi", "Porosity (φ)"), ("sw", "Water Sat (Sw)"),
                            ("ntg", "NTG"), ("bo", "Bo (rb/stb)"), ("bg", "Bg (rcf/scf)")]:
            if key in petro:
                p = petro[key]
                pet_data.append([label, f"{p['p10']:.4f}", f"{p['p50']:.4f}", f"{p['p90']:.4f}", p.get("source","—").upper()])

        t3 = Table(pet_data, colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TEXTCOLOR", (0, 0), (-1, 0), teal),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#162840"), colors.HexColor("#0f2139")]),
            ("TEXTCOLOR", (0, 1), (-1, -1), light),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1a3a4a")),
            ("PADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(t3)

        # Disclaimer
        story.append(Spacer(1, 24))
        story.append(HRFlowable(width="100%", thickness=0.5, color=subtle))
        story.append(Paragraph(
            "Generated by GeoVol Reservoir Volumetrics App. For exploration purposes only. "
            "All results should be verified by a qualified petroleum engineer before use in "
            "investment or development decisions.",
            ParagraphStyle("Disc", parent=styles["Normal"], fontSize=7.5,
                           textColor=subtle, spaceBefore=8)))

        doc.build(story)
        buf.seek(0)
        return buf.getvalue()

    except ImportError:
        return None
    except Exception as e:
        st.error(f"PDF error: {e}")
        return None


def _generate_csv(mc):
    """Generate CSV of Monte Carlo simulation data."""
    dfs = []
    for fluid in ["oiip", "giip"]:
        if fluid in mc:
            r = mc[fluid]
            dfs.append(pd.Series(r["raw"], name=f"{fluid.upper()}_{r['unit']}"))
    if dfs:
        df = pd.concat(dfs, axis=1)
    else:
        df = pd.DataFrame()
    return df.to_csv(index=False)


def _horizon_to_csv(horizon, sample_rate, name):
    """Convert horizon array to CSV (Inline, Crossline, Time_ms)."""
    rows = []
    n_il, n_xl = horizon.shape
    for il in range(n_il):
        for xl in range(n_xl):
            rows.append({"Inline": il, "Crossline": xl,
                         "Time_ms": round(float(horizon[il, xl]) * sample_rate, 1)})
    return pd.DataFrame(rows).to_csv(index=False)
